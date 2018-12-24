# -*- coding: utf-8 -*-
"""
Created on Sat Dec 22 13:47:57 2018

@author: LeeJ6
"""

from urllib.request import urlopen
from bs4 import BeautifulSoup

#stockItem = '092130' # 이크래더블
#stockItem = '034950' # 한국기업평가
stockItem = '007310' # 오뚜기

output_file = stockItem + '.xlsx'
url = 'https://finance.naver.com/item/frgn.nhn?code=' + stockItem
html = urlopen(url)
source = BeautifulSoup(html.read(), 'html.parser')

maxPage=source.find_all("table",align="center")
mp = maxPage[0].find_all("td",class_="pgRR")
mpNum = int(mp[0].a.get('href')[-3:])


# https://code.tutsplus.com/ko/tutorials/how-to-work-with-excel-documents-using-python--cms-25698
# http://pythonstudy.xyz/python/article/405-%ED%8C%8C%EC%9D%B4%EC%8D%AC-%EC%97%91%EC%85%80-%EC%82%AC%EC%9A%A9%ED%95%98%EA%B8%B0
import openpyxl
#excel_document = openpyxl.Workbook()
#excel_document.save('sample.xlsx')
#excel_document = openpyxl.load_workbook('sample.xlsx')
#sheet = excel_document.active
#sheet['A1'] = 100
#sheet.cell(row=2, column=1).value = 200
#excel_document.save('sample.xlsx')

#mpNum = 3
#for page in range(1, mpNum+1):
#    print('page : ', page)
#    url = 'https://finance.naver.com/item/frgn.nhn?code=' + stockItem + '&page=' + str(page)
#    html = urlopen(url)
#    source = BeautifulSoup(html.read(), 'html.parser')
#    srlists = source.find_all('tr', onmouseover='mouseOver(this)')
#    isCheckNone = None
#    
#    
#    for i in range(1, len(srlists) - 1):
#        print(srlists[i].find_all("td", class_="tc")[0].text, srlists[i].find_all("td", class_='num')[0].text, srlists[i].find_all("td", class_='num')[7].text)


import os
import datetime

if os.path.exists(output_file):
    os.remove(output_file)
else:
    print("The file does not exist")

#mpNum = 1
row = 1
excel_document = openpyxl.Workbook()
excel_document.guess_types = True
excel_document.save(output_file)
sheet = excel_document.active

sheet.cell(1, column=1).value = 'Date'
sheet.cell(1, column=2).value = 'Price'
sheet.cell(1, column=3).value = 'Group Buy'
sheet.cell(1, column=4).value = 'Foreign Buy'
sheet.cell(1, column=5).value = 'Foreign Percent'

for page in range(1, mpNum+1):
    print('page : ', page)
    url = 'https://finance.naver.com/item/frgn.nhn?code=' + stockItem + '&page=' + str(page)
    html = urlopen(url)
    source = BeautifulSoup(html.read(), 'html.parser')
    srlists = source.find_all('tr', onmouseover='mouseOver(this)')
    isCheckNone = None
    
    for i in range(1, len(srlists) - 1):
        row = row + 1
        try:
            date_info = srlists[i].find_all("td", class_="tc")[0].text
            cost = srlists[i].find_all("td", class_='num')[0].text
            group_buy = srlists[i].find_all("td", class_='num')[4].text
            foreign_buy = srlists[i].find_all("td", class_='num')[5].text
            foreign_percent = srlists[i].find_all("td", class_='num')[7].text
        except:
            print("Unexpected error] i : ", i, "date_info : ", srlists[i].find_all("td", class_="tc"))
            print("Unexpected error] i : ", i, "field : ", srlists[i].find_all("td", class_='num'))
            break
        else:
            date_info = datetime.datetime.strptime(date_info, '%Y.%m.%d')
            sheet.cell(row, column=1).value = datetime.datetime(date_info.year, date_info.month, date_info.day)
            cost = cost.replace(',', '')
            sheet.cell(row, column=2).value = cost
            group_buy = group_buy.replace('+', '')
            group_buy = group_buy.replace(',', '')
            sheet.cell(row, column=3).value = int(group_buy)
            foreign_buy = foreign_buy.replace('+', '')
            foreign_buy = foreign_buy.replace(',', '')
            sheet.cell(row, column=4).value = int(foreign_buy)
            sheet.cell(row, column=5).value = foreign_percent

excel_document.save(output_file)   
