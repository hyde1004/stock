# -*- coding: utf-8 -*-
"""
Created on Sun Dec 23 15:45:07 2018

@author: LeeJ6
"""

# References
# - 주식 코드 및 주식명 : http://pydata.tistory.com/2
# - csv 읽기 : http://pythonstudy.xyz/python/article/207-CSV-%ED%8C%8C%EC%9D%BC-%EC%82%AC%EC%9A%A9%ED%95%98%EA%B8%B0

import openpyxl
import os
from urllib.request import urlopen
from bs4 import BeautifulSoup

kospi_infile = 'kospi.xlsx'
kospi_outfile = 'kospi_out.xlsx'

kosdaq_infile = 'kosdaq.xlsx'
kosdaq_outfile = 'kosdaq_out.xlsx'

if os.path.exists(kosdaq_outfile):
    os.remove(kosdaq_outfile)
else:
    print("The file does not exist")

excel_document = openpyxl.load_workbook(kosdaq_infile)
#sheet = excel_document.active
sheet = excel_document['kosdaq']
#print(sheet.cell(1, column=1).value)
all_rows = sheet.rows

output_row = 1
excel_document = openpyxl.Workbook()
#excel_document.guess_types = True
excel_document.save(kosdaq_outfile)
sheet = excel_document.active

sheet.cell(output_row, column=1).value = 'Code'
sheet.cell(output_row, column=2).value = 'Name'
sheet.cell(output_row, column=3).value = 'Foreign Percent(Current)'
sheet.cell(output_row, column=4).value = 'Foregin Percent(last month)'
count = 0

for row in all_rows:
    print(row[0].value, row[1].value)
    stockCode = str(row[0].value)
    stockName = row[1].value
    if stockCode == 'Code':
        continue

    #count = count + 1
    #if count > 10:
    #    break
    
    output_row = output_row + 1
    sheet.cell(output_row, column=1).value = stockCode
    sheet.cell(output_row, column=2).value = stockName    


    for page in range(1, 3):
        print('stockCode' + stockCode)
        url = 'https://finance.naver.com/item/frgn.nhn?code=' + stockCode + '&page=' + str(page)
        print('url : ', url)
        html = urlopen(url)
        source = BeautifulSoup(html.read(), 'html.parser')
        srlists = source.find_all('tr', onmouseover='mouseOver(this)')
        isCheckNone = None
        
        i = 0

        try:
            foreign_percent = srlists[i].find_all("td", class_='num')[7].text
        except:
            print("Unexpected error] i : ", i, "date_info : ", srlists[i].find_all("td", class_="tc"))
            print("Unexpected error] i : ", i, "field : ", srlists[i].find_all("td", class_='num'))
            break
        else:
            sheet.cell(output_row, column = page+2).value = foreign_percent

excel_document.save(kosdaq_outfile)           

    

